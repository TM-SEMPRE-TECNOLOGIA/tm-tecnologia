from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from typing import List, Optional
import io
import re
from datetime import date, datetime

from db.supabase import get_supabase
from models.schemas import (
    OrdemServico, OrdemServicoCreate, OrdemServicoUpdate,
    StatusOS, PrioridadeOS, TipoOS, ImportacaoResultado, DashboardKPIs
)

router = APIRouter(prefix="/api/ordens", tags=["ordens"])

TABLE = "ordens_servico"


def _parse_date(value) -> Optional[date]:
    if not value:
        return None
    if isinstance(value, date):
        return value
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return None


# ── CRUD ──────────────────────────────────────────────────

@router.get("", response_model=List[dict])
def listar_ordens(
    contrato_id: Optional[str] = Query(None),
    status: Optional[StatusOS] = Query(None),
    prioridade: Optional[PrioridadeOS] = Query(None),
    tipo: Optional[TipoOS] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
):
    db = get_supabase()
    q = db.table(TABLE).select("*")

    if contrato_id:
        q = q.eq("contrato_id", contrato_id)
    if status:
        q = q.eq("status", status.value)
    if prioridade:
        q = q.eq("prioridade", prioridade.value)
    if tipo:
        q = q.eq("tipo", tipo.value)
    if search:
        q = q.or_(f"numero_os.ilike.%{search}%,titulo.ilike.%{search}%,agencia.ilike.%{search}%")

    offset = (page - 1) * per_page
    q = q.order("created_at", desc=True).range(offset, offset + per_page - 1)

    result = q.execute()
    return result.data or []


@router.get("/kpis", response_model=DashboardKPIs)
def kpis_dashboard(contrato_id: Optional[str] = Query(None)):
    db = get_supabase()
    q = db.table(TABLE).select("status, data_prazo")

    if contrato_id:
        q = q.eq("contrato_id", contrato_id)

    result = q.execute()
    rows = result.data or []

    hoje = date.today()
    total = len(rows)
    abertas = sum(1 for r in rows if r["status"] == "aberta")
    em_andamento = sum(1 for r in rows if r["status"] == "em_andamento")
    concluidas = sum(1 for r in rows if r["status"] == "concluida")
    atrasadas = sum(1 for r in rows if r["status"] == "atrasada")

    vencendo = 0
    for r in rows:
        if r.get("data_prazo") and r["status"] in ("aberta", "em_andamento"):
            try:
                prazo = date.fromisoformat(r["data_prazo"])
                delta = (prazo - hoje).days
                if 0 <= delta <= 7:
                    vencendo += 1
            except (ValueError, TypeError):
                pass

    taxa = round((concluidas / total * 100), 1) if total > 0 else 0.0

    vencendo_30 = 0
    for r in rows:
        if r.get("data_prazo") and r["status"] in ("aberta", "em_andamento"):
            try:
                prazo = date.fromisoformat(r["data_prazo"])
                delta = (prazo - hoje).days
                if 0 <= delta <= 30:
                    vencendo_30 += 1
            except (ValueError, TypeError):
                pass

    # SLA: % concluídas dentro do prazo (tem data_conclusao <= data_prazo)
    com_prazo = [r for r in rows if r.get("data_prazo") and r.get("data_conclusao")]
    no_prazo = 0
    for r in com_prazo:
        try:
            if date.fromisoformat(r["data_conclusao"]) <= date.fromisoformat(r["data_prazo"]):
                no_prazo += 1
        except (ValueError, TypeError):
            pass
    sla_pct = round((no_prazo / len(com_prazo) * 100), 1) if com_prazo else None

    return DashboardKPIs(
        total_os=total,
        abertas=abertas,
        em_andamento=em_andamento,
        concluidas=concluidas,
        atrasadas=atrasadas,
        taxa_conclusao=taxa,
        vencendo_7_dias=vencendo,
        vencendo_30_dias=vencendo_30,
        sla_cumprido_pct=sla_pct,
        contrato_id=contrato_id,
    )


# ── Endpoints de gráficos ─────────────────────────────────

@router.get("/graficos/por-contrato")
def grafico_por_contrato(contrato_id: Optional[str] = Query(None)):
    db = get_supabase()
    q = db.table(TABLE).select("contrato_id, status")
    if contrato_id:
        q = q.eq("contrato_id", contrato_id)
    rows = q.execute().data or []

    contagem: dict[str, dict] = {}
    for r in rows:
        cid = r.get("contrato_id") or "?"
        if cid not in contagem:
            contagem[cid] = {"contrato_id": cid, "total": 0, "concluidas": 0, "abertas": 0}
        contagem[cid]["total"] += 1
        if r["status"] == "concluida":
            contagem[cid]["concluidas"] += 1
        elif r["status"] == "aberta":
            contagem[cid]["abertas"] += 1

    return sorted(contagem.values(), key=lambda x: x["total"], reverse=True)


@router.get("/graficos/evolucao-mensal")
def grafico_evolucao_mensal(contrato_id: Optional[str] = Query(None)):
    from collections import defaultdict
    db = get_supabase()
    q = db.table(TABLE).select("status, data_abertura, data_conclusao, data_prazo")
    if contrato_id:
        q = q.eq("contrato_id", contrato_id)
    rows = q.execute().data or []

    hoje = date.today()
    meses: dict[str, dict] = {}
    # gerar últimos 6 meses
    for i in range(5, -1, -1):
        m = hoje.replace(day=1)
        # subtrair i meses
        month = m.month - i
        year = m.year
        while month <= 0:
            month += 12
            year -= 1
        key = f"{year}-{month:02d}"
        meses[key] = {"mes": key, "abertas": 0, "concluidas": 0}

    for r in rows:
        for campo, bucket in [("data_abertura", "abertas"), ("data_conclusao", "concluidas")]:
            raw = r.get(campo)
            if not raw:
                continue
            try:
                d = date.fromisoformat(str(raw)[:10])
                key = f"{d.year}-{d.month:02d}"
                if key in meses:
                    meses[key][bucket] += 1
            except (ValueError, TypeError):
                pass

    return list(meses.values())


@router.get("/graficos/por-tipo")
def grafico_por_tipo(contrato_id: Optional[str] = Query(None)):
    db = get_supabase()
    q = db.table(TABLE).select("tipo")
    if contrato_id:
        q = q.eq("contrato_id", contrato_id)
    rows = q.execute().data or []

    contagem: dict[str, int] = {}
    for r in rows:
        t = r.get("tipo") or "corretiva"
        contagem[t] = contagem.get(t, 0) + 1

    labels = {"preventiva": "Preventiva", "corretiva": "Corretiva",
               "preditiva": "Preditiva", "inspecao": "Inspeção"}
    return [{"tipo": t, "label": labels.get(t, t), "total": v}
            for t, v in sorted(contagem.items(), key=lambda x: x[1], reverse=True)]


@router.get("/graficos/por-responsavel")
def grafico_por_responsavel(contrato_id: Optional[str] = Query(None)):
    db = get_supabase()
    q = db.table(TABLE).select("responsavel, status")
    if contrato_id:
        q = q.eq("contrato_id", contrato_id)
    rows = q.execute().data or []

    contagem: dict[str, dict] = {}
    for r in rows:
        resp = (r.get("responsavel") or "Sem responsável").strip()
        if not resp:
            resp = "Sem responsável"
        if resp not in contagem:
            contagem[resp] = {"responsavel": resp, "total": 0, "concluidas": 0}
        contagem[resp]["total"] += 1
        if r["status"] == "concluida":
            contagem[resp]["concluidas"] += 1

    top8 = sorted(contagem.values(), key=lambda x: x["total"], reverse=True)[:8]
    return top8


@router.get("/{os_id}", response_model=dict)
def obter_ordem(os_id: int):
    db = get_supabase()
    result = db.table(TABLE).select("*").eq("id", os_id).single().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="O.S não encontrada")
    return result.data


@router.post("", response_model=dict, status_code=201)
def criar_ordem(ordem: OrdemServicoCreate):
    db = get_supabase()
    payload = ordem.model_dump(exclude_none=True)
    # converte date → string ISO
    for k, v in payload.items():
        if isinstance(v, date):
            payload[k] = v.isoformat()
    result = db.table(TABLE).insert(payload).execute()
    return result.data[0]


@router.patch("/{os_id}", response_model=dict)
def atualizar_ordem(os_id: int, dados: OrdemServicoUpdate):
    db = get_supabase()
    payload = dados.model_dump(exclude_none=True)
    for k, v in payload.items():
        if isinstance(v, date):
            payload[k] = v.isoformat()
    if not payload:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    result = db.table(TABLE).update(payload).eq("id", os_id).execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="O.S não encontrada")
    return result.data[0]


@router.delete("/{os_id}", status_code=204)
def deletar_ordem(os_id: int):
    db = get_supabase()
    db.table(TABLE).delete().eq("id", os_id).execute()


# ── Importação Excel ──────────────────────────────────────

# Mapeamento de nomes de colunas (normalizado) → campo canônico
COLUMN_MAP: dict[str, str] = {
    # número OS
    "nro": "numero_os", "nr": "numero_os", "os": "numero_os",
    "numero": "numero_os", "numero_os": "numero_os",
    # título / descrição / serviço
    "titulo": "titulo", "descricao": "descricao",
    "servico": "titulo", "item": "titulo", "servico_descricao": "descricao",
    # status
    "status": "status", "situacao": "status",
    # prioridade / tipo
    "prioridade": "prioridade", "tipo": "tipo",
    # contrato
    "contrato": "contrato_id", "contrato_id": "contrato_id", "lote": "contrato_id",
    # local / agência / prefixo
    "local": "local",
    "agencia": "agencia", "ag": "agencia",
    "prefixo": "agencia_prefixo", "pref": "agencia_prefixo",
    # elaborador (quem elaborou/escreveu o relatório)
    "elaboracao_relatorio": "elaborador", "elaboracao": "elaborador",
    "elaborador": "elaborador", "resp_elaboracao": "elaborador",
    # técnico (quem executa em campo)
    "tecnico": "tecnico", "tecnico_campo": "tecnico",
    # responsável genérico (legacy / fallback)
    "responsavel": "responsavel",
    # datas
    "data_abertura": "data_abertura", "abertura": "data_abertura",
    "data_levant_relatorio": "data_abertura",
    "data_prazo": "data_prazo", "prazo": "data_prazo", "vencimento": "data_prazo",
    "data_conclusao": "data_conclusao", "conclusao": "data_conclusao",
    # valores (armazenamos em observacoes por enquanto)
    "valor_levantamento": "_valor_levant", "valor_aprovado": "_valor_aprov",
    # agendamento / dificuldade → observacoes
    "agendamento": "_agendamento", "dificuldade": "_dificuldade",
}

STATUS_MAP: dict[str, str] = {
    # genéricos
    "aberta": "aberta", "aberto": "aberta", "open": "aberta",
    "em andamento": "em_andamento", "andamento": "em_andamento", "em_andamento": "em_andamento",
    "concluida": "concluida", "concluido": "concluida", "closed": "concluida",
    "atrasada": "atrasada", "atrasado": "atrasada", "late": "atrasada",
    "cancelada": "cancelada", "cancelado": "cancelada",
    # valores reais da planilha TM (já normalizados: sem acento, separadores→espaço)
    "concluida pelo fornecedor": "concluida",
    "orcamento aprovado retorno ao fornecedor": "em_andamento",
    "orcamento aprovado   retorno ao fornecedor": "em_andamento",
    "enviada para orcamento": "em_andamento",
    "em levantamento": "em_andamento",
    "em elaboracao": "em_andamento",
    "em orcamento": "em_andamento",
    "nao quis a preventiva": "cancelada",
    "devolvida": "cancelada",
    "mudanca de contrato": "cancelada",
    "com dificuldade": "em_andamento",
    "fornecedor acionado": "em_andamento",
    "servico concluido pendente relatorio": "concluida",
    "servico concluido  pendente relatorio": "concluida",
    "garantia": "em_andamento",
    "autorizacao pendente": "em_andamento",
    "aguardando autorizacao": "em_andamento",
    "nao autorizado": "cancelada",
}

PRIORIDADE_MAP: dict[str, str] = {
    "baixa": "baixa", "low": "baixa",
    "media": "media", "normal": "media", "medium": "media",
    "alta": "alta", "high": "alta",
    "critica": "critica", "critical": "critica", "urgente": "critica",
}

TIPO_MAP: dict[str, str] = {
    "corretiva": "corretiva", "corretivo": "corretiva",
    "preventiva": "preventiva", "preventivo": "preventiva", "pm": "preventiva", "prev": "preventiva",
    "preditiva": "preditiva", "pdm": "preditiva",
    "inspecao": "inspecao", "inspection": "inspecao",
}

# Extrai contrato_id de strings como "DIVINOPOLIS MG - CTR 2056" → "2056"
_CTR_RE = re.compile(r"(?:CTR|CONTRATO|LOTE)\s+(\d{4})", re.IGNORECASE)
_VENC_SHEET_RE = re.compile(r"\((\d{4})\)")  # ex: "VECIMENTOS MS (6122)"

CONTRATOS_VALIDOS = {"0908","1507","1565","2056","2057","2626","2627","3575","6122"}


def _extract_contrato(valor: str) -> Optional[str]:
    """Extrai o contrato_id de 4 dígitos de uma string descritiva."""
    if not valor:
        return None
    m = _CTR_RE.search(str(valor))
    if m and m.group(1) in CONTRATOS_VALIDOS:
        return m.group(1)
    # tenta encontrar qualquer sequência de 4 dígitos que seja um contrato válido
    for c in CONTRATOS_VALIDOS:
        if c in str(valor):
            return c
    return None


def _normalize_col(col: str) -> str:
    # remove acentos simples e normaliza separadores
    s = str(col).lower().strip()
    s = s.replace("ã","a").replace("á","a").replace("â","a").replace("à","a")
    s = s.replace("é","e").replace("ê","e").replace("è","e")
    s = s.replace("í","i").replace("ï","i")
    s = s.replace("ó","o").replace("ô","o").replace("õ","o")
    s = s.replace("ú","u").replace("ü","u")
    s = s.replace("ç","c").replace("ñ","n")
    return re.sub(r"[\s\-_/]+", "_", s)


def _detect_header_row(ws, max_scan: int = 10) -> int:
    """Detecta a linha do cabeçalho procurando por palavras-chave conhecidas."""
    keywords = {"os","agencia","agência","situacao","situação","vencimento","tecnico","técnico","contrato"}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        cells = {str(c or "").lower().strip() for c in row}
        if len(cells & keywords) >= 2:
            return i + 1  # 1-indexed
    return 1  # fallback: assume linha 1


@router.post("/importar/cabecalhos")
async def extrair_cabecalhos(
    file: UploadFile = File(...),
    aba: Optional[str] = Query(None),
):
    """Extrai cabeçalhos + preview das primeiras 5 linhas + auto-mapeamento sugerido."""
    content = await file.read()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

    abas = [{"nome": n, "linhas": wb[n].max_row or 0} for n in wb.sheetnames]

    header_row = _detect_header_row(ws)
    raw_headers = [str(c.value or "").strip() for c in next(
        ws.iter_rows(min_row=header_row, max_row=header_row)
    )]
    raw_headers = [h for h in raw_headers if h]  # remove vazios

    # Preview: até 5 linhas após o cabeçalho
    preview_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 5, values_only=True):
        preview_rows.append([str(v or "").strip() for v in list(row)[:len(raw_headers)]])

    # Auto-mapeamento sugerido: header original → campo canônico
    sugestoes: dict[str, str] = {}
    for h in raw_headers:
        norm = _normalize_col(h)
        if norm in COLUMN_MAP:
            sugestoes[h] = COLUMN_MAP[norm]

    return {
        "abas": abas,
        "aba_atual": ws.title,
        "cabecalhos": raw_headers,
        "preview": preview_rows,
        "sugestoes": sugestoes,
        "campos_destino": [
            {"campo": "numero_os",    "label": "Número O.S",    "obrigatorio": True},
            {"campo": "agencia",      "label": "Agência",       "obrigatorio": False},
            {"campo": "contrato_id",  "label": "Contrato",      "obrigatorio": False},
            {"campo": "elaborador",   "label": "Elaborador",    "obrigatorio": False},
            {"campo": "tecnico",      "label": "Técnico",       "obrigatorio": False},
            {"campo": "status",       "label": "Situação",      "obrigatorio": False},
            {"campo": "data_prazo",   "label": "Vencimento",    "obrigatorio": False},
            {"campo": "data_abertura","label": "Data Abertura", "obrigatorio": False},
            {"campo": "local",        "label": "Local",         "obrigatorio": False},
            {"campo": "titulo",       "label": "Título/Serviço","obrigatorio": False},
            {"campo": "observacoes",  "label": "Observações",   "obrigatorio": False},
        ],
    }


@router.get("/importar/abas")
async def listar_abas(file: UploadFile = File(...)):
    """Retorna as abas disponíveis no arquivo Excel antes de importar."""
    content = await file.read()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    abas = []
    for name in wb.sheetnames:
        ws = wb[name]
        abas.append({"nome": name, "linhas": ws.max_row or 0})
    return {"abas": abas}


@router.post("/importar", response_model=ImportacaoResultado)
async def importar_excel(
    file: UploadFile = File(...),
    contrato_id: Optional[str] = Query(None),
    aba: Optional[str] = Query(None),
    mapeamento: Optional[str] = Query(None),  # JSON: {"coluna_excel": "campo_destino", ...}
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado. Execute: pip install openpyxl")

    content = await file.read()
    ext = (file.filename or "").lower().split(".")[-1]

    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Formato não suportado. Use .xlsx, .xls ou .csv")

    if ext == "csv":
        import csv
        lines = content.decode("utf-8-sig", errors="replace").splitlines()
        reader = csv.DictReader(lines)
        rows_raw = list(reader)
        headers = list(reader.fieldnames or [])
        header_row = 1
    else:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

        # Detectar linha de cabeçalho
        header_row = _detect_header_row(ws)
        raw_headers = [str(c.value or "").strip() for c in next(
            ws.iter_rows(min_row=header_row, max_row=header_row)
        )]
        rows_raw = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            rows_raw.append(dict(zip(raw_headers, row)))
        headers = raw_headers

    # Mapear colunas — usa mapeamento manual se fornecido, senão auto-detecta
    col_mapping: dict[str, str] = {}
    if mapeamento:
        import json as _json
        try:
            manual: dict = _json.loads(mapeamento)
            # manual = {"coluna_excel": "campo_destino"} — apenas entradas com destino definido
            col_mapping = {k: v for k, v in manual.items() if v and v != "__ignorar__"}
        except Exception:
            pass  # fallback para auto-mapeamento

    if not col_mapping:
        for h in headers:
            norm = _normalize_col(h)
            if norm in COLUMN_MAP:
                col_mapping[h] = COLUMN_MAP[norm]

    # Inferir contrato_id da aba se não foi passado
    sheet_contrato: Optional[str] = None
    if not contrato_id and aba:
        m = _VENC_SHEET_RE.search(aba)
        if m and m.group(1) in CONTRATOS_VALIDOS:
            sheet_contrato = m.group(1)

    erros: list[str] = []
    ordens_para_inserir: list[dict] = []

    for i, row in enumerate(rows_raw, start=header_row + 1):
        try:
            mapped: dict = {}
            extras: list[str] = []  # valores extras → observacoes

            for orig_col, canonical in col_mapping.items():
                val = row.get(orig_col)
                if val is None or str(val).strip() in ("", "None"):
                    continue
                raw_str = str(val).strip()

                if canonical == "status":
                    norm_val = _normalize_col(raw_str).replace("_", " ")
                    mapped["status"] = STATUS_MAP.get(norm_val, STATUS_MAP.get(raw_str.lower(), "aberta"))

                elif canonical == "prioridade":
                    mapped["prioridade"] = PRIORIDADE_MAP.get(raw_str.lower(), "media")

                elif canonical == "tipo":
                    # detecta tipo a partir da descrição se coluna tipo não existir
                    mapped["tipo"] = TIPO_MAP.get(raw_str.lower(), "corretiva")

                elif canonical in ("data_abertura", "data_prazo", "data_conclusao"):
                    parsed = _parse_date(val)
                    if parsed:
                        mapped[canonical] = parsed.isoformat()

                elif canonical == "numero_os":
                    # OS pode vir como float: 250095163.0 → "250095163"
                    try:
                        mapped["numero_os"] = str(int(float(raw_str)))
                    except (ValueError, OverflowError):
                        mapped["numero_os"] = raw_str

                elif canonical == "contrato_id":
                    extracted = _extract_contrato(raw_str)
                    if extracted:
                        mapped["contrato_id"] = extracted

                elif canonical == "agencia_prefixo":
                    mapped["agencia_prefixo"] = raw_str  # tratado depois

                elif canonical.startswith("_"):
                    # campos extras → observacoes
                    label = orig_col
                    extras.append(f"{label}: {raw_str}")

                else:
                    mapped[canonical] = raw_str

            # Juntar extras em observacoes
            if extras:
                obs_existente = mapped.get("observacoes", "")
                joined = " | ".join(extras)
                mapped["observacoes"] = f"{obs_existente} | {joined}".strip(" |") if obs_existente else joined

            # Montar agencia com prefixo se disponível
            prefixo = mapped.pop("agencia_prefixo", None)
            if prefixo and mapped.get("agencia"):
                mapped["agencia"] = f"{mapped['agencia']} [{prefixo}]"
            elif prefixo and not mapped.get("agencia"):
                mapped["agencia"] = prefixo

            # Linha vazia → pula
            if not mapped.get("numero_os") and not mapped.get("agencia") and not mapped.get("titulo"):
                continue

            # Defaults
            if not mapped.get("numero_os"):
                mapped["numero_os"] = f"IMP-{i:04d}"
            if not mapped.get("titulo"):
                mapped["titulo"] = mapped.get("agencia", f"O.S linha {i}")
            if not mapped.get("status"):
                mapped["status"] = "aberta"
            if not mapped.get("prioridade"):
                mapped["prioridade"] = "media"
            if not mapped.get("tipo"):
                mapped["tipo"] = "corretiva"

            # Prioridade de contrato_id: parâmetro > planilha > sheet
            if contrato_id:
                mapped["contrato_id"] = contrato_id
            elif not mapped.get("contrato_id") and sheet_contrato:
                mapped["contrato_id"] = sheet_contrato
            elif not mapped.get("contrato_id"):
                mapped["contrato_id"] = "0000"

            # Remove campos que não existem na tabela
            for campo in list(mapped.keys()):
                if campo not in {
                    "numero_os","titulo","descricao","status","prioridade","tipo",
                    "contrato_id","local","agencia","responsavel",
                    "elaborador","tecnico",
                    "data_abertura","data_prazo","data_conclusao","observacoes"
                }:
                    mapped.pop(campo)

            ordens_para_inserir.append(mapped)
        except Exception as e:
            erros.append(f"Linha {i}: {str(e)[:120]}")

    importadas = 0
    if ordens_para_inserir:
        db = get_supabase()
        for i in range(0, len(ordens_para_inserir), 100):
            batch = ordens_para_inserir[i : i + 100]
            result = db.table(TABLE).insert(batch).execute()
            importadas += len(result.data or [])

    return ImportacaoResultado(
        total_linhas=len(rows_raw),
        importadas=importadas,
        ignoradas=len(rows_raw) - len(ordens_para_inserir),
        erros=erros[:20],
        ordens=ordens_para_inserir[:10],
    )
